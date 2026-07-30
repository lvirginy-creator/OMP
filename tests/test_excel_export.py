import io

import openpyxl
import pytest

from app.models.invoice import Invoice
from app.services import invoice_pipeline
from app.services.excel_export import build_export
from app.services.suppliers import now_iso
from tests.fixtures.pdf_builder import simple_invoice_pdf


def _result(**overrides):
    result = {
        "document_type": "INVOICE",
        "supplier_name": "Fournisseur Export SAS",
        "invoice_number": "F-EXP-1",
        "invoice_date": "2026-04-01",
        "currency": "EUR",
        "total_ht": 58.0,
        "total_vat": 11.6,
        "total_ttc": 69.6,
        "lines": [
            {
                "line_type": "ARTICLE",
                "charge_kind": None,
                "supplier_ref": "EXPREF-1",
                "supplier_label": "Article export",
                "quantity": 5.0,
                "unit_price_net": 10.0,
                "line_total_net": 50.0,
                "vat_rate": 20.0,
                "low_confidence": False,
            },
            {
                "line_type": "CHARGE",
                "charge_kind": "SHIPPING",
                "supplier_ref": None,
                "supplier_label": "Frais de port",
                "quantity": 1.0,
                "unit_price_net": 8.0,
                "line_total_net": 8.0,
                "vat_rate": 20.0,
                "low_confidence": False,
            },
        ],
        "page_count_documents": 1,
    }
    result.update(overrides)
    return result


@pytest.fixture(autouse=True)
def anthropic_key_enabled(monkeypatch):
    from app.core.config import get_settings

    monkeypatch.setenv("ANTHROPIC_API_KEY", "test-key")
    get_settings.cache_clear()
    yield
    get_settings.cache_clear()


def _load(content: bytes) -> openpyxl.Workbook:
    return openpyxl.load_workbook(io.BytesIO(content))


async def _seed_invoice(db_session, monkeypatch, **overrides):
    async def fake_extract_text_mode(text_payload):
        return _result(**overrides)

    monkeypatch.setattr(invoice_pipeline, "extract_text_mode", fake_extract_text_mode)
    result = await invoice_pipeline.process_uploaded_pdf(
        db_session, f"{overrides.get('invoice_number', 'f')}.pdf", simple_invoice_pdf() + str(overrides).encode()
    )
    assert result.outcome == "CREATED"
    return await db_session.get(Invoice, result.invoice_id)


async def test_export_has_two_sheets_with_expected_headers(db_session, monkeypatch):
    await _seed_invoice(db_session, monkeypatch)

    content, filename = await build_export(db_session)
    assert filename.startswith("achats_") and filename.endswith(".xlsx")

    wb = _load(content)
    assert wb.sheetnames == ["Factures", "Détail achats"]
    assert [c.value for c in wb["Factures"][1]] == [
        "Date", "Société", "Fournisseur", "Type", "N° facture", "Total HT", "Total TVA",
        "Total TTC", "dont frais et remises", "Nb lignes articles", "Statut", "Fichier source",
    ]


async def test_charge_lines_excluded_from_detail_but_counted_in_factures(db_session, monkeypatch):
    await _seed_invoice(db_session, monkeypatch)
    content, _ = await build_export(db_session)
    wb = _load(content)

    detail_rows = list(wb["Détail achats"].iter_rows(min_row=2, values_only=True))
    assert len(detail_rows) == 1  # seule la ligne ARTICLE
    assert detail_rows[0][3] == "EXPREF-1"

    factures_rows = list(wb["Factures"].iter_rows(min_row=2, max_row=2, values_only=True))
    assert factures_rows[0][8] == 8.0  # dont frais et remises = la ligne CHARGE


async def test_sum_detail_plus_charges_equals_sum_total_ht(db_session, monkeypatch):
    """§11 : Σ(Montant ligne, onglet 2) + Σ(dont frais et remises, onglet 1) == Σ(Total HT)."""
    await _seed_invoice(db_session, monkeypatch, invoice_number="F-EXP-1")
    await _seed_invoice(
        db_session,
        monkeypatch,
        invoice_number="F-EXP-2",
        invoice_date="2026-04-05",
        total_ht=116.0,
        total_vat=23.2,
        total_ttc=139.2,
        lines=[
            {
                "line_type": "ARTICLE",
                "charge_kind": None,
                "supplier_ref": "EXPREF-2",
                "supplier_label": "Autre article",
                "quantity": 10.0,
                "unit_price_net": 10.0,
                "line_total_net": 100.0,
                "vat_rate": 20.0,
                "low_confidence": False,
            },
            {
                "line_type": "CHARGE",
                "charge_kind": "SHIPPING",
                "supplier_ref": None,
                "supplier_label": "Frais de port",
                "quantity": 1.0,
                "unit_price_net": 16.0,
                "line_total_net": 16.0,
                "vat_rate": 20.0,
                "low_confidence": False,
            },
        ],
    )

    content, _ = await build_export(db_session)
    wb = _load(content)

    detail_sum = sum(
        row[11] for row in wb["Détail achats"].iter_rows(min_row=2, values_only=True) if row[11] is not None
    )
    factures_rows = list(wb["Factures"].iter_rows(min_row=2, values_only=True))
    total_row = factures_rows[-1]
    charges_sum = sum(r[8] for r in factures_rows[:-1] if r[8] is not None)
    total_ht_sum = total_row[5]

    assert round(detail_sum + charges_sum, 2) == round(total_ht_sum, 2)


async def test_unmapped_lines_are_highlighted(db_session, monkeypatch):
    await _seed_invoice(db_session, monkeypatch)
    content, _ = await build_export(db_session)
    wb = _load(content)

    ws = wb["Détail achats"]
    fill = ws.cell(row=2, column=1).fill
    assert fill.fgColor.rgb == "FFFFF2CC"


async def test_mapped_reference_fills_our_ref_and_ean(db_session, monkeypatch):
    from app.models.mapping import Mapping

    invoice = await _seed_invoice(db_session, monkeypatch)
    timestamp = now_iso()
    db_session.add(
        Mapping(
            supplier_id=invoice.supplier_id,
            supplier_ref="EXPREF-1",
            our_ref="INT-EXPREF1",
            our_label="Article interne export",
            ean="1234567890123",
            active=True,
            created_at=timestamp,
            updated_at=timestamp,
        )
    )
    await db_session.commit()

    content, _ = await build_export(db_session)
    wb = _load(content)
    row = list(wb["Détail achats"].iter_rows(min_row=2, max_row=2, values_only=True))[0]
    assert row[5] == "1234567890123"  # code barre
    assert row[7] == "INT-EXPREF1"  # notre référence
    assert row[8] == "Article interne export"


async def test_status_filter_excludes_needs_review(db_session, monkeypatch):
    invoice = await _seed_invoice(db_session, monkeypatch)
    invoice.status = "NEEDS_REVIEW"
    await db_session.commit()

    content, _ = await build_export(db_session, statut="VALIDATED")
    wb = _load(content)
    rows = list(wb["Factures"].iter_rows(min_row=2, values_only=True))
    assert len(rows) == 1  # uniquement la ligne de totaux, aucune facture


async def test_societe_column_and_filter(db_session, monkeypatch):
    from app.models.societe import Societe

    societe_a = Societe(name="OMP A", active=True, created_at=now_iso())
    societe_b = Societe(name="OMP B", active=True, created_at=now_iso())
    db_session.add_all([societe_a, societe_b])
    await db_session.commit()

    async def fake_extract_text_mode(text_payload):
        return _result()

    monkeypatch.setattr(invoice_pipeline, "extract_text_mode", fake_extract_text_mode)
    result_a = await invoice_pipeline.process_uploaded_pdf(
        db_session, "a.pdf", simple_invoice_pdf() + b"A", societe_a.id
    )

    async def fake_extract_text_mode_b(text_payload):
        return _result(invoice_number="F-EXP-2")

    monkeypatch.setattr(invoice_pipeline, "extract_text_mode", fake_extract_text_mode_b)
    result_b = await invoice_pipeline.process_uploaded_pdf(
        db_session, "b.pdf", simple_invoice_pdf() + b"B", societe_b.id
    )
    assert result_a.outcome == "CREATED" and result_b.outcome == "CREATED"

    content, _ = await build_export(db_session)
    wb = _load(content)
    societe_values = {
        row[1] for row in wb["Factures"].iter_rows(min_row=2, values_only=True) if row[1]
    }
    assert societe_values == {"OMP A", "OMP B"}

    content_filtered, _ = await build_export(db_session, societe_ids=[societe_a.id])
    wb_filtered = _load(content_filtered)
    rows = [
        row for row in wb_filtered["Factures"].iter_rows(min_row=2, values_only=True) if row[0]
    ]
    assert len(rows) == 1
    assert rows[0][1] == "OMP A"
