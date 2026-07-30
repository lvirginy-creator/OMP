import io
from datetime import date

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload

from app.models.invoice import Invoice
from app.models.societe import Societe
from app.models.supplier import Supplier
from app.services.mapping_resolution import resolve_mapping

STATUS_LABELS = {"VALIDATED": "Validée", "NEEDS_REVIEW": "À vérifier"}
MONEY_FORMAT = "#,##0.00 €"
DATE_FORMAT = "DD/MM/YYYY"
UNMAPPED_FILL = PatternFill(start_color="FFFFF2CC", end_color="FFFFF2CC", fill_type="solid")


def _fr_date(iso_date: str | None):
    if not iso_date:
        return None
    try:
        return date.fromisoformat(iso_date)
    except ValueError:
        return None


async def _fetch_invoices(
    db: AsyncSession,
    date_du: str | None,
    date_au: str | None,
    supplier_ids: list[int] | None,
    statut: str | None,
    societe_ids: list[int] | None,
) -> list[Invoice]:
    stmt = select(Invoice).options(selectinload(Invoice.lines))
    if date_du:
        stmt = stmt.where(Invoice.invoice_date >= date_du)
    if date_au:
        stmt = stmt.where(Invoice.invoice_date <= date_au)
    if supplier_ids:
        stmt = stmt.where(Invoice.supplier_id.in_(supplier_ids))
    if statut:
        stmt = stmt.where(Invoice.status == statut)
    if societe_ids:
        stmt = stmt.where(Invoice.societe_id.in_(societe_ids))
    return (await db.execute(stmt)).scalars().all()


def _autosize(ws: Worksheet, widths: list[int]) -> None:
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width


async def build_export(
    db: AsyncSession,
    date_du: str | None = None,
    date_au: str | None = None,
    supplier_ids: list[int] | None = None,
    statut: str | None = None,
    societe_ids: list[int] | None = None,
) -> tuple[bytes, str]:
    invoices = await _fetch_invoices(db, date_du, date_au, supplier_ids, statut, societe_ids)

    supplier_names: dict[int, str] = {}
    societe_names: dict[int, str] = {}
    if invoices:
        suppliers = (await db.execute(select(Supplier))).scalars().all()
        supplier_names = {s.id: s.name for s in suppliers}
        societes = (await db.execute(select(Societe))).scalars().all()
        societe_names = {s.id: s.name for s in societes}

    wb = openpyxl.Workbook()

    # --- Onglet 1 : Factures -----------------------------------------
    ws1: Worksheet = wb.active
    ws1.title = "Factures"
    headers1 = [
        "Date", "Société", "Fournisseur", "Type", "N° facture", "Total HT", "Total TVA",
        "Total TTC", "dont frais et remises", "Nb lignes articles", "Statut", "Fichier source",
    ]
    ws1.append(headers1)
    ws1.freeze_panes = "A2"
    ws1.auto_filter.ref = f"A1:{get_column_letter(len(headers1))}1"
    for cell in ws1[1]:
        cell.font = Font(bold=True)

    totals = {"total_ht": 0.0, "total_vat": 0.0, "total_ttc": 0.0, "charges": 0.0}

    sorted_invoices = sorted(
        invoices,
        key=lambda i: (i.invoice_date or "", supplier_names.get(i.supplier_id, ""), i.invoice_number or "", i.id),
    )

    for invoice in sorted_invoices:
        charges = sum(
            (l.line_total_net or 0.0) for l in invoice.lines if l.line_type == "CHARGE"
        )
        nb_articles = sum(1 for l in invoice.lines if l.line_type == "ARTICLE")
        row = [
            _fr_date(invoice.invoice_date),
            societe_names.get(invoice.societe_id, ""),
            supplier_names.get(invoice.supplier_id, ""),
            "Avoir" if invoice.document_type == "CREDIT_NOTE" else "Facture",
            invoice.invoice_number or "",
            invoice.total_ht,
            invoice.total_vat,
            invoice.total_ttc,
            charges,
            nb_articles,
            STATUS_LABELS.get(invoice.status, invoice.status),
            invoice.source_filename,
        ]
        ws1.append(row)
        r = ws1.max_row
        ws1.cell(r, 1).number_format = DATE_FORMAT
        for col in (6, 7, 8, 9):
            ws1.cell(r, col).number_format = MONEY_FORMAT

        totals["total_ht"] += invoice.total_ht or 0.0
        totals["total_vat"] += invoice.total_vat or 0.0
        totals["total_ttc"] += invoice.total_ttc or 0.0
        totals["charges"] += charges

    total_row = ["", "", "", "", "Total", totals["total_ht"], totals["total_vat"], totals["total_ttc"], totals["charges"], "", "", ""]
    ws1.append(total_row)
    r = ws1.max_row
    for col in (5, 6, 7, 8, 9):
        ws1.cell(r, col).font = Font(bold=True)
    for col in (6, 7, 8, 9):
        ws1.cell(r, col).number_format = MONEY_FORMAT

    _autosize(ws1, [12, 20, 28, 10, 16, 12, 12, 12, 18, 14, 12, 28])

    # --- Onglet 2 : Détail achats --------------------------------------
    ws2: Worksheet = wb.create_sheet("Détail achats")
    headers2 = [
        "Date", "Société", "Fournisseur", "Référence fournisseur", "Taille", "Code barre",
        "Libellé fournisseur", "Notre référence", "Notre libellé", "Quantité achetée",
        "Prix unitaire d'achat", "Montant ligne", "N° facture", "Statut facture",
    ]
    ws2.append(headers2)
    ws2.freeze_panes = "A2"
    ws2.auto_filter.ref = f"A1:{get_column_letter(len(headers2))}1"
    for cell in ws2[1]:
        cell.font = Font(bold=True)

    detail_rows = []
    for invoice in invoices:
        for line in invoice.lines:
            if line.line_type != "ARTICLE":
                continue
            detail_rows.append((invoice, line))

    detail_rows.sort(
        key=lambda pair: (
            pair[0].invoice_date or "",
            supplier_names.get(pair[0].supplier_id, ""),
            pair[0].invoice_number or "",
            pair[1].line_no,
        )
    )

    for invoice, line in detail_rows:
        mapping = await resolve_mapping(db, invoice.supplier_id, line.supplier_ref, line.size)
        row = [
            _fr_date(invoice.invoice_date),
            societe_names.get(invoice.societe_id, ""),
            supplier_names.get(invoice.supplier_id, ""),
            line.supplier_ref or "",
            line.size or "",
            mapping.ean if mapping and mapping.ean else "",
            line.supplier_label or "",
            mapping.our_ref if mapping else "",
            mapping.our_label if mapping else "",
            line.quantity,
            line.unit_price_net,
            line.line_total_net,
            invoice.invoice_number or "",
            STATUS_LABELS.get(invoice.status, invoice.status),
        ]
        ws2.append(row)
        r = ws2.max_row
        ws2.cell(r, 1).number_format = DATE_FORMAT
        for col in (11, 12):
            ws2.cell(r, col).number_format = MONEY_FORMAT
        if mapping is None:
            for col in range(1, len(headers2) + 1):
                ws2.cell(r, col).fill = UNMAPPED_FILL

    _autosize(ws2, [12, 20, 28, 16, 10, 16, 28, 16, 28, 12, 14, 12, 16, 12])
    for col_idx in (10, 11, 12):
        for row_idx in range(2, ws2.max_row + 1):
            ws2.cell(row_idx, col_idx).alignment = Alignment(horizontal="right")

    buf = io.BytesIO()
    wb.save(buf)

    all_dates = [i.invoice_date for i in invoices if i.invoice_date] or []
    start = date_du or (min(all_dates) if all_dates else date.today().isoformat())
    end = date_au or (max(all_dates) if all_dates else date.today().isoformat())
    filename = f"achats_{start.replace('-', '')}_{end.replace('-', '')}.xlsx"

    return buf.getvalue(), filename
