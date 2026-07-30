import io
import re
from dataclasses import dataclass, field

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.mapping import Mapping
from app.models.supplier import Supplier
from app.schemas.mapping import MappingRowOutcome
from app.services.normalize import normalize_ref, normalize_size, strip_accents
from app.services.suppliers import get_or_create_supplier, now_iso

TEMPLATE_HEADERS = [
    ("reference_fournisseur", "reference_fournisseur"),
    ("taille", "taille"),
    ("fournisseur", "fournisseur"),
    ("libelle_fournisseur", "libelle_fournisseur"),
    ("notre_reference", "notre_reference"),
    ("notre_libelle", "notre_libelle"),
    ("code_barre", "code_barre"),
]

_HEADER_ALIASES = {
    "reference_fournisseur": "supplier_ref",
    "ref_fournisseur": "supplier_ref",
    "taille": "size",
    "fournisseur": "supplier_name",
    "libelle_fournisseur": "supplier_label",
    "notre_reference": "our_ref",
    "notre_ref": "our_ref",
    "notre_libelle": "our_label",
    "code_barre": "ean",
    "ean": "ean",
    "code_barres": "ean",
}

REQUIRED_FIELDS = ("supplier_ref", "our_ref", "our_label")


def _normalize_header(raw: str) -> str:
    value = strip_accents(str(raw or "")).lower().strip()
    value = re.sub(r"[\s\-']+", "_", value)
    value = re.sub(r"_+", "_", value).strip("_")
    return value


def generate_template() -> bytes:
    wb = openpyxl.Workbook()
    ws: Worksheet = wb.active
    ws.title = "Référentiel"
    ws.append([h for h, _ in TEMPLATE_HEADERS])
    for col_idx in range(1, len(TEMPLATE_HEADERS) + 1):
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = 22
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


@dataclass
class ParsedRow:
    row_number: int
    values: dict[str, str]


@dataclass
class ParseResult:
    rows: list[ParsedRow] = field(default_factory=list)
    header_errors: list[str] = field(default_factory=list)


def parse_workbook(content: bytes) -> ParseResult:
    result = ParseResult()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True, read_only=True)
    except Exception as exc:  # fichier invalide / corrompu
        result.header_errors.append(f"Impossible de lire le fichier Excel : {exc}")
        return result

    ws = wb.worksheets[0]
    rows_iter = ws.iter_rows(values_only=True)
    try:
        header_row = next(rows_iter)
    except StopIteration:
        result.header_errors.append("Le fichier est vide.")
        return result

    field_by_col: dict[int, str] = {}
    for idx, raw_header in enumerate(header_row):
        if raw_header is None:
            continue
        normalized = _normalize_header(raw_header)
        field_name = _HEADER_ALIASES.get(normalized)
        if field_name:
            field_by_col[idx] = field_name

    missing = [f for f in REQUIRED_FIELDS if f not in field_by_col.values()]
    if missing:
        result.header_errors.append(
            "Colonnes obligatoires manquantes : " + ", ".join(missing)
        )
        return result

    for row_number, raw_row in enumerate(rows_iter, start=2):
        if raw_row is None or all(v is None or str(v).strip() == "" for v in raw_row):
            continue
        values: dict[str, str] = {}
        for idx, field_name in field_by_col.items():
            if idx < len(raw_row) and raw_row[idx] is not None:
                values[field_name] = str(raw_row[idx]).strip()
        result.rows.append(ParsedRow(row_number=row_number, values=values))

    return result


async def build_preview(
    db: AsyncSession, parsed: ParseResult
) -> list[MappingRowOutcome]:
    outcomes: list[MappingRowOutcome] = []

    for row in parsed.rows:
        values = row.values
        supplier_ref = values.get("supplier_ref", "").strip()
        our_ref = values.get("our_ref", "").strip()
        our_label = values.get("our_label", "").strip()

        missing = [
            label
            for field_name, label in (
                ("supplier_ref", "reference_fournisseur"),
                ("our_ref", "notre_reference"),
                ("our_label", "notre_libelle"),
            )
            if not values.get(field_name, "").strip()
        ]
        if missing:
            outcomes.append(
                MappingRowOutcome(
                    row_number=row.row_number,
                    supplier_ref=supplier_ref,
                    action="ignore",
                    reason="Champ(s) obligatoire(s) manquant(s) : " + ", ".join(missing),
                )
            )
            continue

        supplier_name = values.get("supplier_name", "").strip() or None
        supplier: Supplier | None = None
        if supplier_name:
            supplier = await get_or_create_supplier(db, supplier_name)

        size_norm = normalize_size(values.get("size"))
        ref_norm = normalize_ref(supplier_ref)
        stmt = select(Mapping).where(Mapping.supplier_ref_norm == ref_norm)
        stmt = stmt.where(Mapping.supplier_id == supplier.id) if supplier else stmt.where(
            Mapping.supplier_id.is_(None)
        )
        stmt = stmt.where(Mapping.size.is_(None)) if size_norm is None else stmt.where(
            Mapping.size == size_norm
        )
        existing = (await db.execute(stmt)).scalar_one_or_none()

        if existing:
            outcomes.append(
                MappingRowOutcome(
                    row_number=row.row_number,
                    supplier_ref=supplier_ref,
                    size=size_norm,
                    supplier_name=supplier_name,
                    our_ref=our_ref,
                    our_label=our_label,
                    action="update",
                    old_our_ref=existing.our_ref,
                    old_our_label=existing.our_label,
                    old_ean=existing.ean,
                )
            )
        else:
            outcomes.append(
                MappingRowOutcome(
                    row_number=row.row_number,
                    supplier_ref=supplier_ref,
                    size=size_norm,
                    supplier_name=supplier_name,
                    our_ref=our_ref,
                    our_label=our_label,
                    action="create",
                )
            )

    return outcomes


async def apply_import(
    db: AsyncSession, parsed: ParseResult
) -> tuple[int, int, int]:
    """Rejoue le parsing et applique l'upsert. Retourne (créés, mis à jour, ignorés)."""
    created = updated = ignored = 0
    timestamp = now_iso()

    for row in parsed.rows:
        values = row.values
        supplier_ref = values.get("supplier_ref", "").strip()
        our_ref = values.get("our_ref", "").strip()
        our_label = values.get("our_label", "").strip()

        if not supplier_ref or not our_ref or not our_label:
            ignored += 1
            continue

        supplier_name = values.get("supplier_name", "").strip() or None
        supplier_label = values.get("supplier_label", "").strip() or None
        ean = values.get("ean", "").strip() or None
        size_norm = normalize_size(values.get("size"))

        supplier: Supplier | None = None
        if supplier_name:
            supplier = await get_or_create_supplier(db, supplier_name)

        ref_norm = normalize_ref(supplier_ref)
        stmt = select(Mapping).where(Mapping.supplier_ref_norm == ref_norm)
        stmt = stmt.where(Mapping.supplier_id == supplier.id) if supplier else stmt.where(
            Mapping.supplier_id.is_(None)
        )
        stmt = stmt.where(Mapping.size.is_(None)) if size_norm is None else stmt.where(
            Mapping.size == size_norm
        )
        existing = (await db.execute(stmt)).scalar_one_or_none()

        if existing:
            existing.our_ref = our_ref
            existing.our_label = our_label
            if supplier_label:
                existing.supplier_label = supplier_label
            if ean:
                existing.ean = ean
            existing.updated_at = timestamp
            updated += 1
        else:
            db.add(
                Mapping(
                    supplier_id=supplier.id if supplier else None,
                    supplier_ref=supplier_ref,
                    size=size_norm,
                    supplier_label=supplier_label,
                    our_ref=our_ref,
                    our_label=our_label,
                    ean=ean,
                    active=True,
                    created_at=timestamp,
                    updated_at=timestamp,
                )
            )
            created += 1

    return created, updated, ignored


async def export_mappings(db: AsyncSession) -> bytes:
    result = await db.execute(
        select(Mapping, Supplier.name)
        .outerjoin(Supplier, Mapping.supplier_id == Supplier.id)
        .order_by(Supplier.name.is_(None), Supplier.name, Mapping.supplier_ref)
    )

    wb = openpyxl.Workbook()
    ws: Worksheet = wb.active
    ws.title = "Référentiel"
    ws.append([h for h, _ in TEMPLATE_HEADERS])

    for mapping, supplier_name in result.all():
        ws.append(
            [
                mapping.supplier_ref,
                mapping.size or "",
                supplier_name or "",
                mapping.supplier_label or "",
                mapping.our_ref,
                mapping.our_label,
                mapping.ean or "",
            ]
        )

    for col_idx in range(1, len(TEMPLATE_HEADERS) + 1):
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = 22

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
